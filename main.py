from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait 
import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from fetcher import fetcher
from fetchermsedc import fetcherms
from fetcherwb import fetcherwb
from fetchertorrentpow import fetchert
from dataloader import load
import polars as pl
from openpyxl import load_workbook, Workbook
import datetime
from selenium.webdriver.common.proxy import Proxy , ProxyType
import ctypes
# from proxy_fetch import get_free_proxies
def loader():
    options = FirefoxOptions()
    #options.add_argument("--headless")
    #options.add_argument("--proxy-server=%s"%myProxy)
    options.page_load_strategy = 'normal'
    
    # webdriver.DesiredCapabilities.FIREFOX['proxy']={
    # 'httpProxy': myProxy,
    # 'sslProxy': myProxy,
    # "socksProxy": myProxy,
    # "socksVersion":4,
    # "proxyType":'manual'
    # } 
    #driver = webdriver.Firefox(options=options)
    driver = webdriver.Firefox(options=options)
    return driver

# def get_free_proxies():
#     options = FirefoxOptions()
#     options.page_load_strategy = "normal"
#     driver = webdriver.Firefox(options=options)
#     driver.get('https://sslproxies.org')

#     table = driver.find_element(By.TAG_NAME, 'table')
#     thead = table.find_element(By.TAG_NAME, 'thead').find_elements(By.TAG_NAME, 'th')
#     tbody = table.find_element(By.TAG_NAME, 'tbody').find_elements(By.TAG_NAME, 'tr')

#     headers = []
#     for th in thead:
#         headers.append(th.text.strip())

#     proxies = []
#     for tr in tbody:
#         proxy_data = {}
#         tds = tr.find_elements(By.TAG_NAME, 'td')
#         for i in range(len(headers)):
#             proxy_data[headers[i]] = tds[i].text.strip()
#         proxies.append(proxy_data)
#     driver.quit()
#     return proxies

def append_to_excel(filename, data):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Provider", "Mobile", "E1", "E2"])  # Add header if file is new

    for row in data:
        sheet.append(row)

    workbook.save(filename)

if __name__=='__main__':  
    ti = datetime.datetime.now().strftime("%Y%m%d%H%M%S") 
    dc,dn,dbu=load()
    df=pl.DataFrame([dc,dn,dbu])
    #print(df)
    #proxies=get_free_proxies()
    # proxy_index=0
    # last_proxy_update_time=time.time()
    
    results = []
    
    iteration_count=0
    
    driver = loader()
    
    for i in range(len(df)):
        # current_time=time.time()
        # if current_time - last_proxy_update_time >=600:
        #     proxies=get_free_proxies()
        #     proxy_index=0
        #     last_proxy_update_time = current_time
        # if iteration_count >= 10:
        #     driver.quit()
        #     proxy_index = (proxy_index + 1)%len(proxies)
        #     driver = loader(proxies[proxy_index]["IP Address"])
        #     iteration_count = 27
        if iteration_count >=50:
            WS_EX_TOPMOST = 0x40000
            windowTitle = "Refresh VPN"
            message = "Click Yes if you have refreshed VPN"

            # display a message box; execution will stop here until user acknowledges
            ctypes.windll.user32.MessageBoxExW(None, message, windowTitle, WS_EX_TOPMOST)

            print("VPN Refreshed")
            iteration_count = 0
        if iteration_count >=20:
            append_to_excel(f"results{ti}.xlsx", results)
            results=[]


        try:
         
            driver.get("https://www.mobikwik.com/electricity-bill-payment")
            WebDriverWait(driver, timeout=15).until(EC.presence_of_element_located((By.XPATH, "//*[@id=\"op\"]/mbk-biller-field/div/div/div[1]/ng-select")))
            
            if(df["Provider"][i]=="mahara"):
                e1,e2=fetcherms(df["Provider"][i],df["Mobile"][i],df["BU"][i],driver)
            
            elif(df["Provider"][i]=="Torrent Power AHMEDABAD" or df["Provider"][i]=="Torrent Power SURAT"):
                last_word = df["Provider"][i].split()[-1]
                operator=df.item(i,"Provider")
                operator=operator.replace(last_word,"")
                operator.rstrip()
                e1,e2=fetchert(operator,df["Mobile"][i],last_word,driver)
                print(operator)
            
            elif(df["Provider"][i]=="west bengal state" or df["Provider"][i]=="dakshin har" or df["Provider"][i]=="UHBVN"):
                e1,e2=fetcherwb(df["Provider"][i],df["Mobile"][i],driver)
            
            else:    
                e1,e2=fetcher(df["Provider"][i],df["Mobile"][i],driver)
            # print(proxies[proxy_index]["IP Address"])
            print(df["Provider"][i],df["Mobile"][i],e1,e2)
            results.append([df["Provider"][i], df["Mobile"][i], e1, e2])
            
            iteration_count = iteration_count + 1
        except Exception as e:
            
            print(f"Error {e}")
            
            #driver.quit()
            #proxies=get_free_proxies()
            # proxy_index = (proxy_index + 1) % len(proxies)
            # driver = loader(proxies[proxy_index]["IP Address"])
            iteration_count = iteration_count + 1
            results.append([df["Provider"][i], df["Mobile"][i], "Main ERROR","Main ERROR"])
            #driver=loader()
            continue   
    append_to_excel(f"results{ti}.xlsx", results)
    print(driver.title)
    driver.quit()