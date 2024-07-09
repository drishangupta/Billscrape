from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait 
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.support import expected_conditions as EC
import polars as pl
from openpyxl import load_workbook, Workbook
import datetime
def loader():
    options = FirefoxOptions()
    options.add_argument("--headless")
    options.page_load_strategy = 'normal'
    driver = webdriver.Firefox(options=options)
    driver.get("https://www.mahadiscom.in/")
    body=driver.find_element(By.TAG_NAME, "body")
    ActionChains(driver).move_to_element(body).send_keys(Keys.TAB).perform()
    ActionChains(driver).move_to_element(body).send_keys(Keys.TAB).perform()
    ActionChains(driver).move_to_element(body).send_keys(Keys.SPACE).perform()
    
    return driver
def append_to_excel(filename, data):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([ "Mobile", "bu"])  # Add header if file is new

    for row in data:
        sheet.append(row)

    workbook.save(filename)
def butake(driver,knumber,i):
    
    body=driver.find_element(By.TAG_NAME, "body")
    original_window = driver.current_window_handle

    WebDriverWait(driver, timeout=2).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#consumerNo")))
    
    text_to_type = f"{knumber}"
    WebDriverWait(driver, timeout=2).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#consumerNo")))
    time.sleep(1)
    driver.find_element(By.CSS_SELECTOR, "#consumerNo").send_keys(Keys.CONTROL,"a")
    driver.find_element(By.CSS_SELECTOR, "#consumerNo").send_keys(Keys.BACK_SPACE)
    
    driver.find_element(By.CSS_SELECTOR, "#consumerNo").send_keys(text_to_type)
    time.sleep(1)
    ActionChains(driver).move_to_element(body).send_keys(Keys.TAB).perform()
    if(i==0):
        ActionChains(driver).move_to_element(body).send_keys(Keys.SPACE).perform()
    ActionChains(driver).move_to_element(body).send_keys(Keys.ENTER).perform()
    time.sleep(1)
    WebDriverWait(driver,timeout=2).until(EC.number_of_windows_to_be(2))
    for window_handle in driver.window_handles:
        if window_handle != original_window:
            driver.switch_to.window(window_handle)
            break
    WebDriverWait(driver,2).until(EC.visibility_of_element_located((By.XPATH,"//*[@id=\"lblBu\"]")))
    bu=driver.find_element(By.XPATH,"//*[@id=\"lblBu\"]")
    bu=bu.text
    driver.close()
    driver.switch_to.window(original_window)

    return bu
    
df = pl.read_excel(source="C://Users//drish//Downloads//MSEDCTEST.xlsx",sheet_name="Sheet1",schema_overrides={"Mobile":pl.String})
knumbers=df["Mobile"]
driver=loader()
results = []
for i in range(len(knumbers)):
    bu=butake(driver,knumbers[i],i)
    print(knumbers[i])
    print(bu)
    results.append([knumbers[i], bu])
ti = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
append_to_excel(f"bu{ti}.xlsx",results)
driver.quit()